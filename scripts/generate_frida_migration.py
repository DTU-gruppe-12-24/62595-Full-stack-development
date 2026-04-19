#!/usr/bin/env python3
import re
import sys
import uuid
import os
import openpyxl

COL_DANISH_NAME  = 0
COL_ENGLISH_NAME = 1
COL_FOOD_ID      = 2
COL_KCAL         = 5
COL_PROTEIN      = 9
COL_CARBS        = 12
COL_FAT          = 14
COL_SALT         = 16
COL_SUGARS       = 97
COL_SAT_FAT      = 180

BATCH_SIZE  = 100
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "R__seed_frida_ingredients.sql")

EXCLUDED_TOP_LEVEL_GROUPS = {126, 134, 149, 195, 196, 197, 198, 199, 200, 215}

RAW_QUALIFIERS = re.compile(
    r",\s*(raw|fresh|dried|frozen|canned|boiled|cooked|roasted|fried|smoked|"
    r"salted|unsalted|unspecified|all varieties|average values?|uspec\.?)\s*$",
    re.IGNORECASE
)

def build_excluded_groups(wb) -> set:
    ws = wb["FoodGroup"]
    group_parent: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        parent_id, group_id = row[0], row[3]
        if group_id is not None:
            group_parent[int(group_id)] = int(parent_id)

    def ancestors(gid: int) -> set:
        result, visited = set(), set()
        while gid in group_parent and gid not in visited:
            visited.add(gid)
            result.add(gid)
            gid = group_parent[gid]
        return result

    excluded = set()
    for gid in group_parent:
        if ancestors(gid) & EXCLUDED_TOP_LEVEL_GROUPS or gid in EXCLUDED_TOP_LEVEL_GROUPS:
            excluded.add(gid)
    return excluded

def build_food_group_map(wb) -> dict:
    ws = wb["Food"]
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        fid, gid = row[2], row[10]
        if fid is not None and gid is not None:
            mapping[int(fid)] = int(gid)
    return mapping

def clean_name(name: str) -> str:
    name = RAW_QUALIFIERS.sub("", name.strip())
    name = re.sub(r"  +", " ", name)
    return name.strip()

def val(row, col):
    v = row[col] if col < len(row) else None
    if v is None:
        return "NULL"
    try:
        return str(round(float(v), 2))
    except (TypeError, ValueError):
        return "NULL"

def escape(s):
    return s.replace("'", "''") if s else ""

def generate(xlsx_path: str):
    print(f"Loading {xlsx_path} ...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    print("Building exclusion list from food groups...")
    excluded_groups = build_excluded_groups(wb)
    food_group_map  = build_food_group_map(wb)

    ws = wb["Data_Table"]
    value_rows = []
    seen_names = set() # NEW: Track unique names
    skipped_groups = 0
    skipped_duplicates = 0

    for row in ws.iter_rows(min_row=5, values_only=True):
        food_id = row[COL_FOOD_ID]
        if food_id is None:
            continue

        group_id = food_group_map.get(int(food_id))
        if group_id in excluded_groups:
            skipped_groups += 1
            continue

        english  = row[COL_ENGLISH_NAME]
        danish   = row[COL_DANISH_NAME]
        raw_name = english if english else danish
        if not raw_name:
            continue

        # Clean and check for duplicates
        display_name = clean_name(raw_name)
        lookup_name = display_name.lower() # Case-insensitive check

        if lookup_name in seen_names:
            skipped_duplicates += 1
            continue

        seen_names.add(lookup_name)

        name    = escape(display_name)
        uid     = str(uuid.uuid4())
        kcal    = val(row, COL_KCAL)
        protein = val(row, COL_PROTEIN)
        carbs   = val(row, COL_CARBS)
        fat     = val(row, COL_FAT)
        salt    = val(row, COL_SALT)
        sugars  = val(row, COL_SUGARS)
        sat_fat = val(row, COL_SAT_FAT)

        value_rows.append(
            f"('{uid}', '{name}', {kcal}, {protein}, {carbs}, {fat}, {sat_fat}, {sugars}, {salt}, NULL)"
        )

    col_list = "(id, name, calories, protein, carbohydrates, fat, saturated_fat, sugars, salt, price)"

    lines = [
        "-- Frida food database seed (auto-generated - do not edit by hand)",
        "-- Re-generate with: python3 generate_frida_migration.py <xlsx>",
        "-- All nutritional values are per 100g/ml",
        "",
        "DELETE FROM ingredients;  -- Clear old data for repeatable migration",
        "",
    ]

    for i in range(0, len(value_rows), BATCH_SIZE):
        chunk = value_rows[i:i + BATCH_SIZE]
        lines.append(f"INSERT INTO ingredients {col_list}")
        lines.append("VALUES")
        for j, row_val in enumerate(chunk):
            comma = "," if j < len(chunk) - 1 else ";"
            lines.append(f"  {row_val}{comma}")
        lines.append("")

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"Skipped {skipped_groups} items due to group exclusion")
    print(f"Skipped {skipped_duplicates} duplicate names")
    print(f"Written {len(value_rows)} unique ingredients -> {OUTPUT_FILE}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python3 generate_frida_migration.py <path-to-xlsx>")
        sys.exit(1)
    generate(sys.argv[1])